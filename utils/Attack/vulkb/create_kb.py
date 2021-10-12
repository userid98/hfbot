from openpyxl import load_workbook
from py2neo import Graph,Node,Relationship,NodeMatcher,RelationshipMatcher
import re
import logging

def nodeExist(lbl, attribute):
    matcher = NodeMatcher(graph)
    
    m = matcher.match(lbl,name=attribute).first()
    
    if m is None:
        #logger.info (lbl,' ',attribute,"not exist")
        return False
    else:
        #logger.info (lbl,' ',attribute," exist")
        return True

def nodeTargetExist(lbl, target,ver):
    matcher = NodeMatcher(graph)
    
    m = matcher.match(lbl,name = target,version = ver).first()
    
    if m is None:
        #logger.info (lbl,' ',attribute,"not exist")
        return False
    else:
        #logger.info (lbl,' ',attribute," exist")
        return True

def relExist(start,end ,rel):
    relmatch=RelationshipMatcher(graph)
    n=relmatch.match([start,end],rel).first()
    
    if n is None:
        #logger.info(start,' ',end,' ',rel,' not exist')
        return False
    else:
        #logger.info(start,' ',end,' ',rel,'  exist')
        return True

def Add_To_neo4j(cve,target,port,os_list,Attack_vector,consequence,cvss_score,access,exp):
    
    #v = Node(VULNERABILITY_LABLE, name=cve,description=Description,vuln_type=Vuln_type ,cvss2_exploitablity= CVSS2_exploitablity,attack_vector= Attack_vector,attack_complexity=Attack_complexity)
    #o = Node(OS_LABLE, name=os)
   # r = Relationship(o, "has", v)

    #s = v | o | r
    #graph.create(s
    
    #插入一个漏洞节点
    if  nodeExist(VULNERABILITY_LABLE,cve):
        a=graph.nodes.match(VULNERABILITY_LABLE,name=cve).first()
    else:
        a = Node(VULNERABILITY_LABLE, name=cve,attack_vector= Attack_vector,cvss_score=cvss_score)
        graph.create(a)
    #插入OS节点
    for i in os_list:
        if  nodeExist(OS_LABLE,i):
            b=graph.nodes.match(OS_LABLE,name=i).first()
        else:
            b = Node(OS_LABLE, name=i)
            graph.create(b)
    #插入关系：vuln affect os
        if not relExist(a,b,VULN_AFFECT_OS) :
            r = Relationship(a, VULN_AFFECT_OS, b)
            graph.create(r)
##########################################################
    # 获取target version list 存入数据库
    tar_name,tar_ver=target_info(target)
    if  nodeTargetExist(TARGET_LABLE,tar_name,tar_ver):
            b = graph.nodes.match(TARGET_LABLE,name=tar_name).first()
    else:
        b = Node(TARGET_LABLE, name=tar_name,ver = tar_ver)
        graph.create(b)
    #插入关系：VULN_ATTACK_TARGET
    if not relExist(a,b,VULN_ATTACK_TARGET) :
        r = Relationship(a, VULN_ATTACK_TARGET, b)
        graph.create(r)

##########################################################
    if  nodeExist(PORT_LABLE,port):
        b=graph.nodes.match(PORT_LABLE,name=port).first()
    else:
        b = Node(PORT_LABLE,name=port)
        graph.create(b)
    #插入关系：VULN_NEED_PORT
    if not relExist(a,b,VULN_NEED_PORT) :
        r = Relationship(a, VULN_NEED_PORT, b)
        graph.create(r)
##########################################################
    # if  nodeExist(POC_LABLE,poc):
    #     b=graph.nodes.match(POC_LABLE,name=poc).first()
    # else:
    #     b = Node(POC_LABLE,name=poc)
    #     graph.create(b)
    # #插入关系：VULN_NEED_PORT
    # if not relExist(a,b,VULN_HAS_POC) :
    #     r = Relationship(a, VULN_HAS_POC, b)
    #     graph.create(r)  
#########################################################
    if  nodeExist(EXP_LABLE,exp):
        b=graph.nodes.match(EXP_LABLE,name=exp).first()
    else:
        b = Node(EXP_LABLE,name=exp)
        graph.create(b)
    #插入关系：VULN_NEED_PORT
    if not relExist(a,b,VULN_HAS_EXP) :
        r = Relationship(a, VULN_HAS_EXP, b)
        graph.create(r)   
#########################################################
    if  nodeExist(CONQUENCE_LABLE,consequence):
        b=graph.nodes.match(CONQUENCE_LABLE,name=consequence).first()
    else:
        b = Node(CONQUENCE_LABLE,name=consequence,access=access)
        graph.create(b)
    #插入关系：VULN_NEED_PORT
    if not relExist(a,b,VULN_CAUSE_IMPACT) :
        r = Relationship(a, VULN_CAUSE_IMPACT, b)
        graph.create(r) 
##########################################################
def target_info(target):
    target_name='init'
    target_ver='no version'
    # 只能提取一个版本号
    # re_str=r'\d+\.(?:\d+\.)*\d+'
    #re_str2=r"\d+(\.\d+)+"
    # target_ver=re.findall(re_str,target)

    
    # 所有规范后的target都能录入
    tar_info = ''
    if target.find(":") != -1:
        tar_info = target.split(":")
        target_ver = tar_info[1]
        target_name = tar_info[0]

        # 规范部分ver的版本
        target_type=['WebLogic','Struts2','Tomcat','WordPress','Nagios','SMBv1','SimpleCalenda','Supervisor XML-RPC server','libssh','Drupal','CouchDB','SquadManagement','vBulletin','ZZZCMS zzzphp','ForgeRock Access Manager','VMware vCenter Server','VMware vRealize-SSRF','RDP','FTP']
        for t in target_type:
            if tar_info[0].find(t)!=-1:
                target_name=t
        if target_name=='init':
            logger.info('target type does not in list')

        target_name = target_name.lower()
    
    return target_name,target_ver

def read_file(address,sheet):
    wb = load_workbook(address)
    logger.info(wb.sheetnames)
    sheet = wb.get_sheet_by_name(sheet)
    max_row=sheet.max_row #最大行数
    for i in range(2,max_row+1):
    
        ## 提取CVE编号、五个属性:description,vuln_type ,CVSS2_exploitablity ,Attack_vector ,Attack_complexity
        cve=sheet.cell(row=i, column=2).value
        target=sheet.cell(row=i, column=3).value
        port=sheet.cell(row=i, column=4).value
        os=sheet.cell(row=i, column=5).value
        Attack_vector=sheet.cell(row=i, column=6).value
        consequence=sheet.cell(row=i, column=7).value
        cvss_score=sheet.cell(row=i, column=8).value
        access=sheet.cell(row=i, column=9).value
        exp=sheet.cell(row=i, column=10).value
        #poc=sheet.cell(row=i, column=11).value
        
        result= [cve,target,port,os,Attack_vector,consequence,cvss_score,access,exp]

        os_list=os.split(';')
        
        Add_To_neo4j(cve,target,str(port),os_list,Attack_vector,consequence,cvss_score,access,exp)
        logger.info(cve)

# graph = Graph(
#     "http://localhost:7474", 
#     username="HFBOT", 
#     password="hfbot"
# )

# log
logger = logging.getLogger(__name__)
logger.setLevel(logging.INFO)
handler = logging.FileHandler("create_kb.log", mode='a')
formatter = logging.Formatter("%(asctime)s -  %(levelname)s - %(message)s")
handler.setFormatter(formatter)
logger.addHandler(handler)
# log

graph = Graph("http://localhost:7474", auth=("neo4j", "hfbot"))
matcher = NodeMatcher(graph)
str1="match (n)-[r]-() delete n,r" 
graph.run(str1)
##节点类型
VULNERABILITY_LABLE="VULNERABILITY"#属性有CVSS_SCORE,ATTACK_VECTOR
OS_LABLE="OS"#
TARGET_LABLE="TARGET"#属性有NAME,VERSION
PORT_LABLE="PORT"#属性有TYPE,ACCESS
#POC_LABLE="POC"#属性有ADDRESS,PARAMETER
EXP_LABLE="EXP"#属性有ADDRESS,PARAMETER
CONQUENCE_LABLE="CONSEQUENCE"#属性有TYPE,ACCESS
##关系类型
VULN_AFFECT_OS="affect"
VULN_ATTACK_TARGET='attack'
VULN_HAS_EXP='has_exp'
#VULN_HAS_POC='has_poc'
VULN_NEED_PORT='need'

VULN_CAUSE_IMPACT='cause'

address="vul_kb.xlsx"
sheet="Sheet1"
read_file(address,sheet)